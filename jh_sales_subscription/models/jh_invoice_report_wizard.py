from odoo import models, fields, api, _
from odoo.exceptions import UserError
import io
import base64
from datetime import datetime, date
import logging

_logger = logging.getLogger(__name__)


class jh_invoice_report_wizard(models.TransientModel):
    _name = 'jh.invoice.report.wizard'
    _description = 'Wizard para Exportar Reporte de Facturas'

    # Filtros de fecha
    date_from = fields.Date(
        string='Fecha Desde',
        required=True,
        default=lambda self: fields.Date.context_today(self).replace(day=1)
    )
    date_to = fields.Date(
        string='Fecha Hasta',
        required=True,
        default=fields.Date.context_today
    )

    # Filtros adicionales
    partner_ids = fields.Many2many(
        'res.partner',
        string='Clientes',
        domain=[('customer_rank', '>', 0)]
    )

    company_ids = fields.Many2many(
        'res.company',
        string='Compañías',
        default=lambda self: self.env.company
    )

    product_ids = fields.Many2many(
        'product.product',
        string='Productos'
    )

    state = fields.Selection([
        ('draft', 'Borrador'),
        ('posted', 'Publicado'),
        ('cancel', 'Cancelado')
    ], string='Estado')

    move_type = fields.Selection([
        ('out_invoice', 'Factura de Cliente'),
        ('out_refund', 'Nota de Crédito Cliente'),
        ('in_invoice', 'Factura de Proveedor'),
        ('in_refund', 'Nota de Crédito Proveedor')
    ], string='Tipo de Documento')

    agents_ids = fields.Many2many(
        'res.partner',
        'jh_invoice_report_wizard_agent_rel',
        'jh_invoice_report_wizard_id',
        'partner_id',
        string='Agentes',
        domain=[('agent', '=', True)]
    )

    # Selección de columnas - CORREGIDAS según campos reales
    show_move = fields.Boolean(string='Número de Factura', default=True)
    show_partner = fields.Boolean(string='Cliente/Proveedor', default=True)
    show_commercial_partner = fields.Boolean(string='Contacto Principal', default=False)
    show_invoice_date = fields.Boolean(string='Fecha Factura', default=True)
    show_invoice_date_due = fields.Boolean(string='Fecha Vencimiento', default=False)
    show_product = fields.Boolean(string='Producto', default=True)
    show_product_categ = fields.Boolean(string='Categoría Producto', default=False)
    show_quantity = fields.Boolean(string='Cantidad', default=True)
    show_product_uom = fields.Boolean(string='Unidad de Medida', default=False)
    show_price_average = fields.Boolean(string='Precio Promedio', default=True)
    show_price_subtotal = fields.Boolean(string='Subtotal', default=True)
    show_price_total = fields.Boolean(string='Total', default=True)
    show_price_margin = fields.Boolean(string='Margen', default=False)
    show_currency = fields.Boolean(string='Moneda', default=True)
    show_invoice_user = fields.Boolean(string='Vendedor', default=True)
    show_team = fields.Boolean(string='Equipo de Ventas', default=False)
    show_company = fields.Boolean(string='Compañía', default=False)
    show_country = fields.Boolean(string='País', default=False)
    show_journal = fields.Boolean(string='Diario', default=False)
    show_payment_state = fields.Boolean(string='Estado de Pago', default=False)
    show_fiscal_position = fields.Boolean(string='Posición Fiscal', default=False)
    show_account = fields.Boolean(string='Cuenta Contable', default=False)
    show_partner_shipping = fields.Boolean(string='Dirección de Entrega', default=False)

    # Campos personalizados JH
    show_jh_commission = fields.Boolean(string='Comisión', default=False)
    show_jh_commission_percent = fields.Boolean(string='% Comisión', default=True)
    show_jh_margin_percent = fields.Boolean(string='% Margen', default=False)
    show_jh_tax_amount = fields.Boolean(string='Valor IVA', default=False)
    show_jh_cost = fields.Boolean(string='Costo', default=False)
    show_jh_agents = fields.Boolean(string='Agentes', default=False)
    show_jh_serial_number = fields.Boolean(string='N° de Serie', default=True)
    show_jh_payment_mode = fields.Boolean(string='Forma de Pago', default=True)
    show_jh_payment_term = fields.Boolean(string='Condiciones de Pago', default=True)

    # Formato de exportación
    export_format = fields.Selection([
        ('xlsx', 'Excel (.xlsx)'),
        ('csv', 'CSV (.csv)')
    ], string='Formato', default='xlsx', required=True)

    # Archivo generado
    file_data = fields.Binary(string='Archivo', readonly=True)
    file_name = fields.Char(string='Nombre del Archivo', readonly=True)
    state_wizard = fields.Selection([
        ('draft', 'Configuración'),
        ('done', 'Completado')
    ], default='draft')

    @api.constrains('date_from', 'date_to')
    def _check_dates(self):
        for record in self:
            if record.date_from > record.date_to:
                raise UserError(_('La fecha desde no puede ser mayor a la fecha hasta.'))

    def _get_domain(self):
        """Construir el dominio basado en los filtros seleccionados"""
        domain = [
            ('invoice_date', '>=', self.date_from),
            ('invoice_date', '<=', self.date_to),
        ]

        if self.partner_ids:
            domain.append(('partner_id', 'in', self.partner_ids.ids))

        if self.company_ids:
            domain.append(('company_id', 'in', self.company_ids.ids))

        if self.product_ids:
            domain.append(('product_id', 'in', self.product_ids.ids))

        if self.state:
            domain.append(('state', '=', self.state))

        if self.move_type:
            domain.append(('move_type', '=', self.move_type))

        if self.agents_ids:
            agent_names = self.agents_ids.mapped('name')
            domain += ['|'] * (len(agent_names) - 1)
            domain += [('jh_agents', 'ilike', name) for name in agent_names]

        return domain

    def _get_selected_fields(self):
        """Obtener los campos seleccionados para exportar"""
        # Mapeo de campos REALES del modelo account.invoice.report
        fields_dict = {}

        if self.show_move:
            fields_dict['move_id'] = 'Número de Factura'
        if self.show_partner:
            fields_dict['partner_id'] = 'Cliente/Proveedor'
        if self.show_commercial_partner:
            fields_dict['commercial_partner_id'] = 'Contacto Principal'
        if self.show_invoice_date:
            fields_dict['invoice_date'] = 'Fecha Factura'
        if self.show_invoice_date_due:
            fields_dict['invoice_date_due'] = 'Fecha Vencimiento'
        if self.show_product:
            fields_dict['product_id'] = 'Producto'
        if self.show_product_categ:
            fields_dict['product_categ_id'] = 'Categoría Producto'
        if self.show_quantity:
            fields_dict['quantity'] = 'Cantidad'
        if self.show_product_uom:
            fields_dict['product_uom_id'] = 'Unidad de Medida'
        if self.show_price_average:
            fields_dict['price_average'] = 'Precio Promedio'
        if self.show_price_subtotal:
            fields_dict['price_subtotal'] = 'Subtotal'
        if self.show_price_total:
            fields_dict['price_total'] = 'Total'
        if self.show_price_margin:
            fields_dict['price_margin'] = 'Margen'
        if self.show_currency:
            fields_dict['currency_id'] = 'Moneda'
        if self.show_invoice_user:
            fields_dict['invoice_user_id'] = 'Vendedor'
        if self.show_partner_shipping:
            fields_dict['jh_partner_shipping_id'] = 'Dirección de Entrega'
        if self.show_team:
            fields_dict['team_id'] = 'Equipo de Ventas'
        if self.show_company:
            fields_dict['company_id'] = 'Compañía'
        if self.show_country:
            fields_dict['country_id'] = 'País'
        if self.show_journal:
            fields_dict['journal_id'] = 'Diario'
        if self.show_payment_state:
            fields_dict['payment_state'] = 'Estado de Pago'
        if self.show_fiscal_position:
            fields_dict['fiscal_position_id'] = 'Posición Fiscal'
        if self.show_account:
            fields_dict['account_id'] = 'Cuenta Contable'

        # Campos personalizados JH
        if self.show_jh_commission:
            fields_dict['jh_commission'] = 'Comisión'
        if self.show_jh_commission_percent:
            fields_dict['jh_commission_percent'] = '% Comisión'
        if self.show_jh_margin_percent:
            fields_dict['jh_margin_percent'] = '% Margen'
        if self.show_jh_tax_amount:
            fields_dict['jh_tax_amount'] = 'Valor IVA'
        if self.show_jh_cost:
            fields_dict['jh_cost'] = 'Costo'
        if self.show_jh_agents:
            fields_dict['jh_agents'] = 'Agentes'
        if self.show_jh_serial_number:
            fields_dict['jh_serial_number'] = 'N° de Serie'
        if self.show_jh_payment_mode:
            fields_dict['jh_payment_mode_id'] = 'Forma de Pago'
        if self.show_jh_payment_term:
            fields_dict['jh_payment_term_id'] = 'Condiciones de Pago'

        return fields_dict

    def _get_serial_number_explicit(self, record):
        """
        Calcular el número de serie de forma explícita para account.invoice.report
        Replica la lógica de commission.settlement.line._compute_serial_number
        pero adaptada para account.invoice.report
        """
        try:
            if not record.move_id or not record.product_id:
                return ''

            # Buscar la línea de factura correspondiente
            invoice_line = self.env['account.move.line'].search([
                ('move_id', '=', record.move_id.id),
                ('product_id', '=', record.product_id.id)
            ], limit=1)

            if not invoice_line:
                return ''

            # Méodo 1: Buscar directamente en jh_sale_lot_id de la línea de factura
            try:
                if hasattr(invoice_line, 'jh_sale_lot_id') and invoice_line.jh_sale_lot_id:
                    sale_lot_value = invoice_line.jh_sale_lot_id
                    if isinstance(sale_lot_value, models.Model):
                        serial = sale_lot_value.name or str(sale_lot_value) or ''
                    else:
                        serial = str(sale_lot_value).strip()
                    if serial:
                        return serial
            except Exception:
                pass

            # Méodo 2: Buscar lot_id directamente en la línea de factura
            try:
                if hasattr(invoice_line, 'lot_id') and invoice_line.lot_id:
                    serial = invoice_line.lot_id.name or ''
                    if serial:
                        return serial
            except Exception:
                pass

            # Méodo 3: Buscar en la línea de venta vinculada
            sale_line = False
            try:
                if hasattr(invoice_line, 'sale_line_ids') and invoice_line.sale_line_ids:
                    sale_line = invoice_line.sale_line_ids[0]
                else:
                    sale_line = self.env['sale.order.line'].search([
                        ('invoice_lines', 'in', invoice_line.id)
                    ], limit=1)
            except Exception:
                pass

            if sale_line:
                # Méodo 3a: Buscar lot_id directamente en la línea de venta
                try:
                    if hasattr(sale_line, 'lot_id') and sale_line.lot_id:
                        serial = sale_line.lot_id.name or ''
                        if serial:
                            return serial
                except Exception:
                    pass

                # Méodo 3b: Buscar en movimientos de stock
                try:
                    domain = [('sale_line_id', '=', sale_line.id)]
                    if invoice_line.product_id:
                        domain.append(('product_id', '=', invoice_line.product_id.id))
                    stock_moves = self.env['stock.move'].search(domain)

                    if stock_moves:
                        move_lines = stock_moves.mapped('move_line_ids')
                        if move_lines:
                            lot_names = move_lines.mapped('lot_id.name')
                            serial = ', '.join(filter(None, lot_names))
                            if not serial:
                                lot_names = move_lines.mapped('lot_name')
                                serial = ', '.join(filter(None, filter(lambda x: x, lot_names)))
                            if serial:
                                return serial
                except Exception:
                    pass

                # Méodo 3c: Búsqueda directa en stock.move.line
                try:
                    domain = [('move_id.sale_line_id', '=', sale_line.id)]
                    if invoice_line.product_id:
                        domain.append(('product_id', '=', invoice_line.product_id.id))
                    move_lines_direct = self.env['stock.move.line'].search(domain)

                    if move_lines_direct:
                        lot_names = move_lines_direct.mapped('lot_id.name')
                        serial = ', '.join(filter(None, lot_names))
                        if not serial:
                            lot_names = move_lines_direct.mapped('lot_name')
                            serial = ', '.join(filter(None, filter(lambda x: x, lot_names)))
                        if serial:
                            return serial
                except Exception:
                    pass

            # Méodo 4: Buscar en pickings relacionados
            try:
                sale_orders = invoice_line.move_id.invoice_line_ids.mapped('sale_line_ids.order_id')
                if sale_orders:
                    pickings = self.env['stock.picking'].search([
                        ('sale_id', 'in', sale_orders.ids),
                        ('state', '=', 'done')
                    ])

                    if pickings:
                        all_move_lines = pickings.mapped('move_line_ids_without_package')
                        if invoice_line.product_id:
                            picking_move_lines = all_move_lines.filtered(
                                lambda ml: ml.product_id.id == invoice_line.product_id.id
                            )
                        else:
                            picking_move_lines = all_move_lines

                        if picking_move_lines:
                            lot_names = picking_move_lines.mapped('lot_id.name')
                            serial = ', '.join(filter(None, lot_names))
                            if not serial:
                                lot_names = picking_move_lines.mapped('lot_name')
                                serial = ', '.join(filter(None, filter(lambda x: x, lot_names)))
                            if serial:
                                return serial
            except Exception:
                pass

            return ''
        except Exception:
            return ''

    def _get_field_value(self, record, field_name):
        """Obtener el valor de un campo de forma segura"""
        try:
            if not hasattr(record, field_name):
                return ''

            # Caso especial para jh_serial_number: forzar cálculo explícito
            if field_name == 'jh_serial_number':
                return self._get_serial_number_explicit(record)

            value = getattr(record, field_name, False)

            # Formatear valores según el tipo
            if isinstance(value, models.Model):
                # Para move_id, queremos el nombre/número de la factura
                if field_name == 'move_id' and value:
                    return value.name or ''
                return value.display_name if value else ''
            elif isinstance(value, bool):
                return 'Sí' if value else 'No'
            elif isinstance(value, (int, float)):
                return value
            elif isinstance(value, datetime):
                return value.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(value, date):
                return value.strftime('%Y-%m-%d')
            else:
                return str(value) if value else ''
        except Exception as e:
            return ''

    def _prepare_data(self, records, fields_dict):
        """Preparar los datos para exportar"""
        data = []

        for record in records:
            row = {}
            for field_name, field_label in fields_dict.items():
                row[field_label] = self._get_field_value(record, field_name)
            data.append(row)

        return data

    def _export_xlsx(self, data, fields_dict):
        """Exportar a formato Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError(_('Por favor instala la librería openpyxl: pip install openpyxl'))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte Facturas'

        # Estilos
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Escribir encabezados
        headers = list(fields_dict.values())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Escribir datos
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
                cell.border = thin_border

                # Alinear números a la derecha
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                    # Formato de número con 2 decimales
                    cell.number_format = '#,##0.00'

        # Ajustar ancho de columnas
        for idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(idx)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 12)

        # Congelar primera fila
        ws.freeze_panes = 'A2'

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return base64.b64encode(output.read())

    def _export_csv(self, data, fields_dict):
        """Exportar a formato CSV"""
        import csv

        output = io.StringIO()
        headers = list(fields_dict.values())

        writer = csv.DictWriter(output, fieldnames=headers, delimiter=';')
        writer.writeheader()
        writer.writerows(data)

        return base64.b64encode(output.getvalue().encode('utf-8-sig'))

    def action_generate_report(self):
        """Generar el reporte"""
        self.ensure_one()

        # Obtener dominio y campos
        domain = self._get_domain()
        fields_dict = self._get_selected_fields()

        if not fields_dict:
            raise UserError(_('Debe seleccionar al menos una columna para exportar.'))

        # Buscar registros con search() normal, no search_read()
        invoice_report = self.env['account.invoice.report'].sudo()
        records = invoice_report.search(domain)

        if not records:
            raise UserError(_('No se encontraron registros con los filtros seleccionados.'))

        # Preparar datos
        data = self._prepare_data(records, fields_dict)

        # Generar archivo según formato
        if self.export_format == 'xlsx':
            file_data = self._export_xlsx(data, fields_dict)
            file_name = f'reporte_facturas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        else:
            file_data = self._export_csv(data, fields_dict)
            file_name = f'reporte_facturas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        # Actualizar wizard
        self.write({
            'file_data': file_data,
            'file_name': file_name,
            'state_wizard': 'done'
        })

        # Mensaje de éxito
        message = _('Se han exportado %s registros correctamente.') % len(data)

        # Retornar acción para mantener el wizard abierto
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'jh.invoice.report.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
            'context': dict(self.env.context, message=message),
        }

    def action_back(self):
        """Volver a la configuración"""
        self.state_wizard = 'draft'
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'jh.invoice.report.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
